import ast
import base64
import csv
import io
import logging
import json
import os
import re
import traceback

from datetime import date, datetime
from collections import defaultdict
from io import BytesIO, StringIO
from typing import Any, Optional
from ..ozon_api import get_posting_from_ozon_api

from ..helpers import (
    split_list,
    split_keywords,
    split_keywords_on_slash,
    remove_latin_characters,
    remove_duplicates_from_list,
    mean,
)

import openpyxl
from odoo import models, fields, api, exceptions
from odoo.exceptions import UserError

from ..ozon_api import (
    ALL_COMMISSIONS,
    get_product_info_list_by_sku,
)
from ..helpers import convert_ozon_datetime_str_to_odoo_datetime_str


logger = logging.getLogger()


class ImportFile(models.Model):
    _name = "ozon.import_file"
    _description = "Импорт"
    _order = "timestamp desc"

    timestamp = fields.Date(
        string="Дата импорта", default=fields.Date.today, readonly=True
    )
    model = fields.Char(string="К какой модели относиться график")

    data_for_download = fields.Selection(
        [
            ("ozon_products", "Товары Ozon"),
            ("ozon_plugin", "Товары Ozon (Плагин)"),
            ("ozon_commissions", "Комиссии Ozon по категориям"),
            ("ozon_transactions", "Транзакции Ozon"),
            ("ozon_stocks", "Остатки товаров Ozon"),
            ("ozon_prices", "Цены Ozon"),
            ("ozon_images", "Ссылки на графики"),
            ("ozon_successful_products_competitors", "Успешные товары конкурентов"),
            ("ozon_postings", "Отправления Ozon"),
            ("ozon_fbo_supply_orders", "Поставки FBO"),
            ("ozon_actions", "Акции Ozon"),
            ("ozon_competitors_goods", "Товары ближайших конкурентов Ozon с продажами"),
            (
                "ozon_ad_campgaign_search_promotion_report",
                "Отчёт о рекламной кампании (продвижение в поиске)",
            ),
            ("ozon_realisation_report", "Отчёт о реализации товаров"),
        ],
        string="Данные для загрузки",
    )

    file = fields.Binary(
        attachment=True, string="Файл для загрузки своих данных", help="Выбрать файл"
    )
    period_start_date = fields.Date(string="Период с")
    period_end_date = fields.Date(string="по")

    def name_get(self):
        """
        Rename name records
        """
        result = []
        for record in self:
            id = record.id
            name = f"Загруженный файл № {id}"
            result.append((id, name))
        return result
    
    @staticmethod
    def get_content(file_value):
        content = base64.b64decode(file_value)
        try:
            content = content.decode("utf-8")
        except UnicodeDecodeError:
            logger.error(f"content.decode error: {traceback.format_exc()}")
        return content

    @api.model
    def create(self, values):
        self.import_file(values)
        return super(ImportFile, self).create(values)

    def import_file(self, values) -> dict:
        log_data = {}
        if not "file" in values or not values["file"]:
            raise exceptions.ValidationError("Отсутствует файл.")

        elif not "data_for_download" in values or not values["data_for_download"]:
            raise exceptions.ValidationError("Необходимо выбрать 'Данные для загрузки'")

        content = self.get_content(file_value=values["file"])

        if values["data_for_download"] == "ozon_competitors_goods":
            workbook = openpyxl.load_workbook(BytesIO(base64.b64decode(values["file"])))
            self._parse_ozon_competitors_goods_xlsx_file(workbook)

        elif values["data_for_download"] == "ozon_images":
            if values["model"] == "sale":
                self.import_images_sale(content)

            elif values["model"] == "sale_by_week":
                self.import_images_sale_by_week(content)

            elif values["model"] == "competitors_products":
                self.import_images_competitors_products(content)

            elif values["model"] == "price_history":
                self.import_images_price_history(content)

            elif values["model"] == "stock":
                self.import_images_stock(content)

            elif values["model"] == "analysis_data":
                self.import_images_analysis_data(content)

            elif values["model"] == "categorie_analysis_data":
                self.import_images_categorie_analysis_data(content)

            elif values["model"] == "categorie_sale_this_year":
                self.import_images_categorie_categorie_sale_this_year(content)

            elif values["model"] == "categorie_sale_last_year":
                self.import_images_categorie_categorie_sale_last_year(content)

        elif values["data_for_download"] == "ozon_successful_products_competitors":
            self.import_successful_products_competitors(content)

        elif values["data_for_download"] == "ozon_ad_campgaign_search_promotion_report":
            start_date, end_date = self.import_ad_campgaign_search_promotion_report(content)
            values['period_start_date'] = start_date
            values['period_end_date'] = end_date

        elif values["data_for_download"] == "ozon_realisation_report":
            self.import_ozon_realisation_report(content)

        elif values["data_for_download"] == "ozon_products":
            log_data = self.process_products_imported_data(content)

        elif values["data_for_download"] == "ozon_commissions":
            with StringIO(content) as csvfile:
                reader = csv.DictReader(csvfile)
                self.env["ozon.ozon_fee"].search([]).unlink()
                cats = {i["name_categories"]: i["id"] 
                        for i in self.env["ozon.categories"].search([]).read(["name_categories"])}
                data = []
                for i, row in enumerate(reader):
                    name = row["commission_name"]
                    cat_id = row["description_category_id"]
                    cat_name = row["category_name"]
                    if cat_id := cats.get(cat_name):
                        data.append(
                            {
                                "name": name,
                                "value": row["value"],
                                "category": cat_id,
                                "type": row["commission_type"],
                                "trading_scheme": row["trading_scheme"],
                            }
                        )
                        print(f"{i}th commission for category {cat_name} was created")
                self.env["ozon.ozon_fee"].create(data)

        elif values["data_for_download"] == "ozon_transactions":
            log_data = self.import_transactions(content)

        elif values["data_for_download"] == "ozon_stocks":
            log_data = self.import_stocks(content)

        elif values["data_for_download"] == "ozon_prices":
            log_data = self.import_prices(content)

        elif values["data_for_download"] == "ozon_postings":
            log_data = self.import_postings(content)

        elif values["data_for_download"] == "ozon_fbo_supply_orders":
            log_data = self.import_fbo_supply_orders(content)

        elif values["data_for_download"] == "ozon_actions":
            log_data = self.import_actions(content)

        return log_data

    def is_ozon_product_exists(self, id_on_platform: str):
        result = self.env["ozon.products"].search(
            [("id_on_platform", "=", id_on_platform)],
            limit=1,
        )
        return result if result else False

    def is_ozon_product_exists_by_sku(self, sku: str):
        result = self.env["ozon.products"].search(
            ["|", "|", ("sku", "=", sku), ("fbo_sku", "=", sku), ("fbs_sku", "=", sku)],
            limit=1,
        )
        return result if result else False


    def is_ozon_product_exists_by_sku_or_article(self, sku: str, article: str):
        result = self.env["ozon.products"].search(
            ["|", "|", "|", ("sku", "=", sku), ("fbo_sku", "=", sku), ("fbs_sku", "=", sku), ("article", "=", article)],
            limit=1,
        )
        return result if result else False

    def is_ozon_category_exists_by_name(self, category_name):
        result = self.env["ozon.categories"].search(
            [("name_categories", "=", category_name)],
            limit=1,
        )
        return result if result else False

    def is_ozon_category_exists_by_id(self, category_id):
        result = self.env["ozon.categories"].search([("c_id", "=", category_id)])
        return result if result else False

    def is_retail_product_exists(self, product_id):
        result = self.env["retail.products"].search(
            [("product_id", "=", product_id)],
            limit=1,
        )
        return result if result else False

    def is_retail_seller_exists(self, seller_name):
        result = self.env["retail.seller"].search(
            [("name", "=", seller_name)],
            limit=1,
        )
        return result if result else False

    def is_ozon_fee_exists(self, category_name, commission_name):
        result = self.env["ozon.ozon_fee"].search(
            [
                ("category.name_categories", "=", category_name),
                ("name", "=", commission_name),
            ],
            limit=1,
        )
        return result if result else False

    def is_ozon_transaction_exists(self, transaction_id):
        result = self.env["ozon.transaction"].search(
            [("transaction_id", "=", transaction_id)],
            limit=1,
        )
        return result if result else False

    def is_ozon_service_exists(self, service_name):
        result = self.env["ozon.ozon_services"].search(
            [("name", "=", service_name)],
            limit=1,
        )
        return result if result else False

    def get_posting(self, posting_number: str):
        posting = self.env["ozon.posting"].search([
            ('posting_number', '=', posting_number)
        ])
        if not posting:
            posting = self.env["ozon.posting"].search([
                ('order_number', '=', posting_number)
            ], order='create_date desc', limit=1)
            if not posting:
                try:
                    posting_data = get_posting_from_ozon_api(posting_number)
                    if posting_data:
                        self.import_postings(posting_data)

                        posting = self.env["ozon.posting"].search([
                            ('posting_number', '=', posting_number)
                        ])
                        if not posting:
                            posting = self.env["ozon.posting"].search([
                                ('order_number', '=', posting_number)
                            ], order='create_date desc', limit=1)
                except Exception:
                    logger.warning(f"Exception get_posting: {traceback.format_exc()}")

        return posting

    @staticmethod
    def get_vals_to_create_value_by_product(posting, row, service_list, posting_number):
        vals_to_create_value_by_product = []
        qty_new_decomposed_transactions = 0
        by_volume = {
            'логистика',
            'магистраль',
            'обратная логистика',
            'обратная магистраль',
            'обработка невыкупа',
            'обработка отмен',
            'обработка возврата',
            'MarketplaceServiceItemRedistributionReturnsPVZ',
            'сборка заказа',
            'обработка отправления'
        }
        by_price = {
            'последняя миля',
        }
        if posting:
            tran_accruals_for_sale = float(row["accruals_for_sale"])
            tran_sale_commission = float(row["sale_commission"])
            tran_amount = float(row["amount"])
            posting_products = posting.posting_product_ids
            products_t_volume = sum(
                product.ozon_products_id.products.volume * product.quantity for product in posting_products
            )
            products_t_price = sum(product.price * product.quantity for product in posting_products)
            for product in posting_products:
                qty = product.quantity
                product_price = product.price
                product_id = product.ozon_products_id.id
                product_volume = product.ozon_products_id.products.volume
                for _ in range(qty):
                    if tran_accruals_for_sale:
                        accruals_for_sale = product_price if tran_accruals_for_sale >= 0 else -product_price
                        sale_commission = abs((product_price * tran_sale_commission) / tran_accruals_for_sale)
                        if tran_sale_commission < 0:
                            sale_commission = -sale_commission
                        amount = (product_price * tran_amount) / tran_accruals_for_sale
                        if tran_amount < 0:
                            amount = -amount if amount >= 0 else amount
                    else:
                        tran_accruals_for_sale = products_t_price
                        accruals_for_sale = product_price
                        sale_commission = abs((product_price * tran_sale_commission) / tran_accruals_for_sale)
                        if tran_sale_commission < 0:
                            sale_commission = -sale_commission
                        amount = (product_price * tran_amount) / tran_accruals_for_sale
                        if tran_amount < 0:
                            amount = -amount if amount >= 0 else amount

                    if row['name'] in [
                        "Доставка покупателю",
                    ]:
                        vals_to_create_value_by_product.extend([
                            {
                                "name": "Сумма за заказы",
                                "value": accruals_for_sale,
                                "ozon_products_id": product_id
                            },
                            {
                                "name": "Вознаграждение за продажу",
                                "value": sale_commission,
                                "ozon_products_id": product_id
                            },
                            {
                                "name": "Итого за заказы",
                                "value": amount,
                                "ozon_products_id": product_id
                            }
                        ])
                        qty_new_decomposed_transactions += 3
                        # {'обработка отправления', 'сборка заказа',
                        # 'последняя миля', 'логистика', 'магистраль'}
                        for name, price in service_list:
                            if name in by_volume:
                                service_amount = (product_volume * price) / products_t_volume
                            else:
                                service_amount = (product_price * price) / tran_accruals_for_sale
                            vals_to_create_value_by_product.append({
                                "name": name,
                                "value": service_amount,
                                "ozon_products_id": product_id
                            })
                            qty_new_decomposed_transactions += 1
                    elif row['name'] in [
                        "Получение возврата, отмены, невыкупа от покупателя"
                    ]:
                        vals_to_create_value_by_product.extend([
                            {
                                "name": "Получение возврата: Сумма за заказы",
                                "value": accruals_for_sale,
                                "ozon_products_id": product_id
                            },
                            {
                                "name": "Получение возврата: Вознаграждение за продажу",
                                "value": sale_commission,
                                "ozon_products_id": product_id
                            },
                            {
                                "name": "Получение возврата: Итого за заказы",
                                "value": amount,
                                "ozon_products_id": product_id
                            }
                        ])
                        qty_new_decomposed_transactions += 3
                        # {'обработка отправления', 'сборка заказа',
                        # 'последняя миля', 'логистика', 'магистраль'}
                        for name, price in service_list:
                            if name in by_volume:
                                service_amount = (product_volume * price) / products_t_volume
                            else:
                                service_amount = (product_price * price) / tran_accruals_for_sale
                            vals_to_create_value_by_product.append({
                                "name": name,
                                "value": service_amount,
                                "ozon_products_id": product_id
                            })
                            qty_new_decomposed_transactions += 1
                    elif row['name'] in [
                        "Доставка и обработка возврата, отмены, невыкупа",
                        "Доставка покупателю — отмена начисления"
                    ]:
                        # {'обработка невыкупа', 'обработка отмен', 'обработка возврата',
                        # 'MarketplaceServiceItemRedistributionReturnsPVZ', 'последняя миля',
                        # 'логистика', 'обратная логистика', 'обратная магистраль', 'магистраль'}
                        for name, price in service_list:
                            if name in by_volume:
                                service_amount = (product_volume * price) / products_t_volume
                            else:
                                service_amount = (product_price * price) / tran_accruals_for_sale
                            vals_to_create_value_by_product.append({
                                "name": name,
                                "value": service_amount,
                                "ozon_products_id": product_id
                            })
                            qty_new_decomposed_transactions += 1
                    else:
                        # оплата эквайринга, услуга продвижения «Бонусы продавца»,
                        vals_to_create_value_by_product.append({
                            "name": row["name"],
                            "value": amount,
                            "ozon_products_id": product_id
                        })
                        qty_new_decomposed_transactions += 1

        else:
            if row['name'] in [
                'Услуги продвижения товаров',
                'Услуга размещения товаров на складе',
                'Приобретение отзывов на платформе',
                'Обработка отправления «Pick-up» (отгрузка курьеру)',
                'Начисления по претензиям',
                'Начисление по спору',
                'Утилизация',
            ]:
                vals_to_create_value_by_product.append({
                    "name": row["name"],
                    "value": row["amount"]
                })
                qty_new_decomposed_transactions += 1
            else:
                vals_to_create_value_by_product.append({
                    "name": row["name"],
                    "value": row["amount"]
                })
                qty_new_decomposed_transactions += 1
                logger.warning(f"posting_number: {posting_number}")
                logger.warning(row["name"])
                logger.warning("Импорт транзакций: не найдено отправление. "
                               "Декомпозированая транзакция создана без разделения на товары.")

        return vals_to_create_value_by_product, qty_new_decomposed_transactions

    def import_transactions(self, content) -> dict:
        qty_new_transactions = 0
        qty_new_decomposed_transactions = 0
        imported_days_data = defaultdict(int)
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            all_transactions = (
                self.env["ozon.transaction"]
                .search([])
                .mapped(lambda r: r.transaction_id)
            )
            all_transactions = dict.fromkeys(all_transactions, True)
            transactions_data = []
            for i, row in enumerate(reader):
                imported_days_data[row["transaction_date"]] += 1
                if all_transactions.get(row["transaction_id"]):
                    logger.warning(f"{i} - Transaction already exists")
                    continue

                # services
                service_list = ast.literal_eval(row["services"])
                services_cost = 0
                services_vals_to_create = []
                for name, price in service_list:
                    services_vals_to_create.append({"name": name, "price": price})
                    services_cost += price

                # posting
                posting_number = row["posting_number"]
                posting = self.get_posting(posting_number)

                decomposed_transactions = self.get_vals_to_create_value_by_product(
                    posting=posting,
                    row=row,
                    service_list=service_list,
                    posting_number=posting_number
                )
                vals_to_create_value_by_product, qty = decomposed_transactions
                qty_new_decomposed_transactions += qty

                ozon_products = []
                ozon_products_ids = []
                skus = ast.literal_eval(row["product_skus"])
                for sku in skus:
                    if ozon_product := self.is_ozon_product_exists_by_sku(sku):
                        ozon_products.append(ozon_product)
                        ozon_products_ids.append(ozon_product.id)

                # TODO: как быть с транзакциями, где указанных sku нет в нашей системе?
                # или напр. в транзакции 2 sku, один из них есть у нас, другого нет.
                # if len(skus) != len(ozon_products):
                #     print(f"{i} -")
                #     continue

                tran_data = {
                    "transaction_id": str(row["transaction_id"]),
                    "transaction_date": row["transaction_date"],
                    "order_date": row["order_date"],
                    "name": row["name"],
                    "accruals_for_sale": row["accruals_for_sale"],
                    "sale_commission": row["sale_commission"],
                    "transaction_type": row["type"],
                    "amount": row["amount"],
                    "skus": skus,
                    "products": ozon_products_ids,
                    "services": [(0, 0, vals) for vals in services_vals_to_create],
                    "posting_number": row["posting_number"],
                    "ozon_transaction_value_by_product_ids": [(0, 0, vals) for vals in vals_to_create_value_by_product]
                }
                transactions_data.append(tran_data)

                qty_new_transactions += 1
                print(f"{i} - Transaction {row['transaction_id']} was created")
                # creating ozon.sale records
                if row["name"] == "Доставка покупателю":
                    self.create_sale_from_transaction(data=tran_data, 
                        products=ozon_products, services_cost=services_cost)
            self.env["ozon.transaction"].create(transactions_data)

        log_data = {
            'Новых транзакций импортировано': qty_new_transactions,
            'Декомпозированых транзакций создано': qty_new_transactions,
            'Записей получено по датам': ''}
        for date_, qty in imported_days_data.items():
            log_data[date_] = qty

        return log_data

    def get_all_products_dict_by_id_on_platform(self) -> dict:
        query = """
                           SELECT
                               id_on_platform,
                               id,
                               stocks_fbs,
                               stocks_fbo
                           FROM ozon_products
                           """
        self.env.cr.execute(query)
        products_raw_vals = self.env.cr.fetchall()
        products_dict = {product[0]: {
            'id': product[1],
            'stocks_fbs': product[2],
            'stocks_fbo': product[3],
        } for product in products_raw_vals}

        return products_dict

    name = fields.Char(string="Название")
    w_id = fields.Char(string="Идентификатор")

    def get_all_warehouses_data(self):
        query = """
                                   SELECT
                                       w_id,
                                       id
                                   FROM ozon_warehouse
                                   """
        self.env.cr.execute(query)
        warehouse_raw_vals = self.env.cr.fetchall()
        warehouses = {}
        for wh in warehouse_raw_vals:
            if not warehouses.get(int(wh[0])):
                warehouses[int(wh[0])] = wh[1]

        return warehouses

    def import_stocks(self, content):
        all_products_by_id_on_platform = self.get_all_products_dict_by_id_on_platform()
        stocks_vals = []
        stocks_imported_data = {}
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                id_on_platform = row["id_on_platform"]
                if ozon_product := all_products_by_id_on_platform.get(id_on_platform):
                    if ozon_product:
                        ozon_product_id = ozon_product['id']
                        vals = {
                            "product": ozon_product_id,
                            "id_on_platform": id_on_platform,
                            "stocks_fbs": row["stocks_fbs"],
                            "stocks_fbo": row["stocks_fbo"],
                        }
                        stocks_vals.append(vals)
                        stocks_imported_data[id_on_platform] = {
                            "stocks_fbs": row["stocks_fbs"],
                            "stocks_fbo": row["stocks_fbo"],
                            "id_on_platform": row["id_on_platform"],
                            "sku": row["sku"],
                            "stocks_fbs_warehouses": row["stocks_fbs_warehouses"],
                            "ozon_product_id": ozon_product_id,
                        }

        stocks = self.env["ozon.stock"].create(stocks_vals)
        all_warehouses_data = self.get_all_warehouses_data()
        fbs_warehouse_product_stock_vals_to_write = []
        products_ids_and_stocks_to_write = {}
        products_ids = []
        for stock in stocks:
            id_on_platform = stock.id_on_platform
            row = stocks_imported_data[id_on_platform]
            product_data = all_products_by_id_on_platform[id_on_platform]

            stocks_by_warehouse = ast.literal_eval(row["stocks_fbs_warehouses"])

            for item in stocks_by_warehouse:
                warehouse_id = all_warehouses_data.get(item["warehouse_id"])
                if not warehouse_id:
                    warehouse_id = self.env["ozon.warehouse"].create(
                        {"name": item["warehouse_name"], "w_id": item["warehouse_id"]}
                    ).id
                    logger.info(f"Создан новый warehouse с id {warehouse_id}")
                    all_warehouses_data = self.get_all_warehouses_data()

                qty = item["present"]
                fbs_warehouse_product_stock_vals_to_write.append(
                    {
                        "stock_id": stock.id,
                        "product_id": row['ozon_product_id'],
                        "warehouse_id": warehouse_id,
                        "qty": qty,
                    }
                )

            # product qty
            new_stocks_fbs = int(row["stocks_fbs"])
            new_stocks_fbo = int(row["stocks_fbo"])
            if (
                    product_data['stocks_fbs'] != new_stocks_fbs or
                    product_data['stocks_fbo'] != new_stocks_fbo
            ):
                product_id = product_data['id']
                stocks = {
                    "stocks_fbs": new_stocks_fbs,
                    "stocks_fbo": new_stocks_fbo,
                }
                products_ids_and_stocks_to_write[product_id] = stocks
                products_ids.append(product_id)

        st_records = self.env["ozon.fbs_warehouse_product_stock"].create(fbs_warehouse_product_stock_vals_to_write)
        logger.warning(f"Создано записей остатков на FBS складах {len(st_records)}")

        updated_products_qty = self._write_stocks_to_products(products_ids, products_ids_and_stocks_to_write)
        logger.warning(f"Обновлено остатков товаров {updated_products_qty}")

        log_data = {
            "Обновлено остатков товаров": updated_products_qty,
            "Создано записей остатков на FBS складах": len(st_records),
        }

        return log_data

    def _write_stocks_to_products(self, products_ids: list, products_ids_and_stocks_to_write: dict[dict]) -> int:
        ozon_products = self.env['ozon.products'].browse(products_ids)
        for product in ozon_products:
            vals = products_ids_and_stocks_to_write[product.id]
            product.write(vals)

        return len(ozon_products)

    def import_prices(self, content) -> dict:
        prices_qty = 0
        updated_products_qty = 0
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                prices_qty += 1
                if ozon_product := self.is_ozon_product_exists(
                    id_on_platform=row["id_on_platform"]
                ):
                    ozon_product.write(
                        {
                            "price": row["price"],
                            "old_price": row["old_price"],
                            "ext_comp_min_price": row["ext_comp_min_price"],
                            "ozon_comp_min_price": row["ozon_comp_min_price"],
                            "self_marketplaces_min_price": row[
                                "self_marketplaces_min_price"
                            ],
                            "price_index": row["price_index"],
                        }
                    )
                    updated_products_qty += 1

        log_data = {
            'Получено данных о ценах': prices_qty,
            'Цены обновлены для количества продуктов': updated_products_qty
        }

        return log_data

    def get_or_create_warehouse(self, warehouse_id, warehouse_name):
        """Returns existing warehouse or create a new one"""
        if warehouse := self.env["ozon.warehouse"].search(
            [("w_id", "=", warehouse_id)], limit=1
        ):
            pass
        else:
            warehouse = self.env["ozon.warehouse"].create(
                {"name": warehouse_name, "w_id": warehouse_id}
            )
        return warehouse

    def import_postings(self, content):
        qty_postings = 0
        qty_created = 0
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            data = []
            for i, row in enumerate(reader):
                qty_postings += 1
                posting_number = row["posting_number"]
                status = row["status"]
                if self.env["ozon.posting"].search(
                    [("posting_number", "=", posting_number), ("status", "=", status)]
                ):
                    continue
                """
                Создаем отправление только если хотя бы один из товаров в отправлении
                соответствует нашему товару
                """
                products = ast.literal_eval(row["products"])
                # [{'offer_id': '063478', 'price': '650.0000', 'quantity': 1, 'sku': 415273036}]
                skus = []
                product_ids = []
                ozon_posting_product_vals_to_create = []
                for product_data in products:
                    sku = str(product_data['sku'])
                    skus.append(sku)
                    article = product_data['offer_id']
                    if ozon_product := self.is_ozon_product_exists_by_sku_or_article(sku, article):
                        product_ids.append(ozon_product.id)

                    ozon_product_id = ozon_product.id if ozon_product else False
                    ozon_posting_product_vals_to_create.append({
                        "ozon_products_id": ozon_product_id,
                        "offer_id": article,
                        "price": float(product_data['price']),
                        "quantity": int(product_data['quantity']),
                        "sku": sku,
                    })
                if not product_ids:
                    continue

                warehouse = self.get_or_create_warehouse(
                    warehouse_id=row["warehouse_id"],
                    warehouse_name=row["warehouse_name"],
                )

                data.append(
                    {
                        "posting_number": posting_number,
                        "in_process_at": row["in_process_at"],
                        "trading_scheme": row["trading_scheme"],
                        "order_id": row["order_id"],
                        "order_number": row["order_number"],
                        "status": status,
                        "product_ids": product_ids,
                        "skus": skus,
                        "region": row["region"],
                        "city": row["city"],
                        "warehouse_id": warehouse.id,
                        "cluster_from": row["cluster_from"],
                        "cluster_to": row["cluster_to"],
                        "posting_product_ids": [(0, 0, vals) for vals in ozon_posting_product_vals_to_create]
                    }
                )
                qty_created += 1
                # print(f"{i} - Posting {row['posting_number']} was imported")

            self.env["ozon.posting"].create(data)

        log_data = {
            "Получено данных об отправлениях": qty_postings,
            "Отправлений создано": qty_created,
        }

        return log_data

    def import_fbo_supply_orders(self, content) -> dict:
        qty_fbo_supply_order = 0
        qty_created = 0
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            for i, row in enumerate(reader):
                qty_fbo_supply_order += 1
                supply_order_id = row["supply_order_id"]
                if self.env["ozon.fbo_supply_order"].search(
                    [("supply_order_id", "=", supply_order_id)]
                ):
                    continue
                warehouse = self.get_or_create_warehouse(
                    warehouse_id=row["warehouse_id"],
                    warehouse_name=row["warehouse_name"],
                )
                fbo_supply_order_data = {
                    "created_at": row["created_at"],
                    "supply_order_id": supply_order_id,
                    "total_items_count": row["total_items_count"],
                    "total_quantity": row["total_quantity"],
                    "warehouse_id": warehouse.id,
                }
                if row["supply_date"]:
                    fbo_supply_order_data["supply_date"] = row["supply_date"]

                fbo_supply_order = self.env["ozon.fbo_supply_order"].create(
                    fbo_supply_order_data
                )
                items = ast.literal_eval(row["items"])
                fbo_supply_order_product_data = []
                skus = []
                for item in items:
                    sku = item["sku"]
                    if ozon_product := self.is_ozon_product_exists_by_sku(sku):
                        fbo_supply_order_product_data.append(
                            {
                                "fbo_supply_order_id": fbo_supply_order.id,
                                "product_id": ozon_product.id,
                                "qty": item["qty"],
                            }
                        )
                        skus.append(sku)
                fbo_supply_order_products = self.env[
                    "ozon.fbo_supply_order_product"
                ].create(fbo_supply_order_product_data)
                fbo_supply_order.write(
                    {
                        "fbo_supply_order_products_ids": fbo_supply_order_products.ids,
                        "skus": skus,
                    }
                )
                qty_created += 1
                # print(f"{i} - Supply order {supply_order_id} was imported")

        log_data = {
            "Данных о заказах на поставку получено": qty_fbo_supply_order,
            "Создано новых заказов на поставку": qty_created,
        }

        return log_data

    def create_sale_from_transaction(self, data: dict, products: list, services_cost: float):
        if len(products) == 0:
            return
        # if all products are the same
        product_id = data["products"][0]
        qty = len(data["products"])
        if data["products"].count(product_id) == qty:
            self.env["ozon.sale"].create(
                {
                    "transaction_identifier": data["transaction_id"],
                    "product": product_id, 
                    "product_id_on_platform": products[0].id_on_platform, 
                    "date": data["order_date"], 
                    "qty": qty, 
                    "revenue": data["accruals_for_sale"],
                    "sale_commission": data["sale_commission"],
                    "services_cost": services_cost,
                    "profit": data["amount"]
                }
            )
        # TODO: что делать если разные продукты в одной транзакции? как создавать продажи?

    def import_ad_campgaign_search_promotion_report(self, content) -> tuple:
        f_path = "/mnt/extra-addons/ozon/__pycache__/ad_campgaign_search_promotion_report.csv"
        with open(f_path, "w") as f:
            f.write(content)
        with open(f_path) as csvfile:
            first_row = next(csvfile)
            ad_campaign = first_row[first_row.find("№") + 2 : first_row.find(",")]
            if not ad_campaign.isdigit():
                raise UserError(
                    """Файл должен содержать номер рекламной кампании в первой строке"""
                )

            date_pattern = r"\b\d{2}\.\d{2}\.\d{4}\b"
            dates = re.findall(date_pattern, first_row)
            period_dates = [datetime.strptime(date_, "%d.%m.%Y") for date_ in dates]
            if len(dates) == 2 and isinstance(period_dates[0], date) and isinstance(period_dates[1], date):
                start_date = period_dates[0] if period_dates[0] < period_dates[1] else period_dates[1]
                end_date = period_dates[1] if period_dates[1] > period_dates[0] else period_dates[0]
            else:
                raise UserError("Файл должен содержать даты начала и конца периода в первой строке"
                                " в формате 01.01.2024-05.02.2024")

            reader = csv.DictReader(csvfile, delimiter=";")
            if reader.fieldnames != [
                "Дата",
                "ID заказа",
                "Номер заказа",
                "Ozon ID",
                "Ozon ID продвигаемого товара",
                "Артикул",
                "Наименование",
                "Количество",
                "Цена продажи",
                "Стоимость, ₽",
                "Ставка, %",
                "Ставка, ₽",
                "Расход, ₽",
            ]:
                raise UserError(
                    """Неправильный файл. Файл должен содержать столбцы:\n'Дата', 'ID заказа', 'Номер заказа', 'Ozon ID', 'Ozon ID продвигаемого товара', 'Артикул', 'Наименование', 'Количество', 'Цена продажи', 'Стоимость, ₽', 'Ставка, %', 'Ставка, ₽', 'Расход, ₽'"""
                )
            promo_type = "search"
            data = []
            for i, row in enumerate(reader):
                sku = row["Ozon ID продвигаемого товара"]
                order_id = row["ID заказа"]
                if self.env["ozon.promotion_expenses"].search(
                    [
                        ("ad_campaign", "=", ad_campaign),
                        ("order_id", "=", order_id),
                    ]
                ):
                    continue
                if ozon_product := self.is_ozon_product_exists_by_sku(sku):
                    if (
                        posting_ids := self.env["ozon.posting"]
                        .search(
                            [
                                ("order_id", "=", order_id),
                            ]
                        )
                        .ids
                    ):
                        pass
                    else:
                        posting_ids = None
                    data.append(
                        {
                            "ad_campaign": ad_campaign,
                            "date": datetime.strptime(row["Дата"], "%d.%m.%Y").date(),
                            "promotion_type": promo_type,
                            "product_id": ozon_product.id,
                            "sku": sku,
                            "order_id": order_id,
                            "posting_ids": posting_ids,
                            "price": row["Цена продажи"].replace(",", "."),
                            "qty": row["Количество"],
                            "total_price": row["Стоимость, ₽"].replace(",", "."),
                            "percent_rate": row["Ставка, %"].replace(",", "."),
                            "abs_rate": row["Ставка, ₽"].replace(",", "."),
                            "expense": row["Расход, ₽"].replace(",", "."),
                        }
                    )
                    print(f"{i} - Product (SKU: {sku}) promotion expenses were added")
            self.env["ozon.promotion_expenses"].create(data)
        os.remove(f_path)

        return start_date, end_date

    def import_ozon_realisation_report(self, content):
        with StringIO(content) as jsonfile:
            data = json.load(jsonfile)
            if self.env["ozon.realisation_report"].search([("num", "=", data["header"]["num"])]):
                print(f"""Report №{data["header"]["num"]} already exists""")
                return
            report = self.env["ozon.realisation_report"].create({**data["header"]})
            products_in_report_data = []
            for i, row in enumerate(data["rows"]):
                product_id_on_platform = row.pop("product_id")
                product = self.is_ozon_product_exists(product_id_on_platform)
                if product:
                    products_in_report_data.append({
                        "product_id": product.id,
                        "product_id_on_platform": product_id_on_platform,
                        "realisation_report_id": report.id,
                        **row
                    })
                    print(f"{i} - Product {product_id_on_platform} added to realisation report")
            self.env["ozon.realisation_report_product"].create(products_in_report_data)

    # def import_images_sale(self, content):
    #     model_products = self.env["ozon.products"]
    #
    #     (
    #         product_id,
    #         url_this_year,
    #         url_last_year,
    #         data_this_year,
    #         data_last_year,
    #     ) = content.split(",")
    #
    #     record = model_products.search([("id", "=", product_id)])
    #
    #     record.img_url_sale_this_year = url_this_year
    #     record.img_url_sale_last_year = url_last_year
    #
    #     record.img_data_sale_this_year = data_this_year.replace("|", ",")
    #     record.img_data_sale_last_year = data_last_year.replace("|", ",")

    # def import_images_sale_by_week(self, content):
    #     model_products = self.env["ozon.products"]
    #
    #     (
    #         product_id,
    #         url_two_weeks,
    #         url_six_weeks,
    #         url_twelve_weeks,
    #         data_two_weeks,
    #         data_six_week,
    #         data_twelve_week,
    #     ) = content.split(",")
    #
    #     record = model_products.search([("id", "=", product_id)])
    #
    #     record.img_url_sale_two_weeks = url_two_weeks
    #     record.img_url_sale_six_weeks = url_six_weeks
    #     record.img_url_sale_twelve_weeks = url_twelve_weeks
    #
    #     record.img_data_sale_two_weeks = data_two_weeks.replace("|", ",")
    #     record.img_data_sale_six_weeks = data_six_week.replace("|", ",")
    #     record.img_data_sale_twelve_weeks = data_twelve_week.replace("|", ",")

    # def import_images_competitors_products(self, content):
    #     model_competitors_products = self.env["ozon.products_competitors"]
    #
    #     product_id, url, data = content.split(",")
    #
    #     record = model_competitors_products.search([("id", "=", product_id)])
    #     record.imgs_url_this_year = url
    #     record.imgs_data_graph_this_year = data.replace("|", ",")

    # def import_images_price_history(self, content):
    #     model_products = self.env["ozon.products"]
    #
    #     product_id, url, data = content.split(",")
    #
    #     record = model_products.search([("id", "=", product_id)])
    #
    #     record.img_url_price_history = url
    #     record.img_data_price_history = data.replace("|", ",")

    # def import_images_stock(self, content):
    #     model_products = self.env["ozon.products"]
    #
    #     product_id, url, data = content.split(",")
    #
    #     record = model_products.search([("id", "=", product_id)])
    #
    #     record.img_url_stock = url
    #     record.img_data_stock = data.replace("|", ",")

    # def import_images_analysis_data(self, content):
    #     model_products = self.env["ozon.products"]
    #
    #     product_id, url, hits_view, hits_tocart = content.split(",")
    #
    #     record = model_products.search([("id", "=", product_id)])
    #
    #     record.img_url_analysis_data = url
    #     record.img_data_analysis_data = {
    #         "hits_view": hits_view,
    #         "hits_tocart": hits_tocart,
    #     }

    # def import_images_categorie_analysis_data(self, content):
    #     model_categories = self.env["ozon.categories"]
    #
    #     model, categories_id, url, data_hits, data_tocart = content.split(",")
    #     data_hits = data_hits.replace("|", ",").replace("'", '"')
    #     data_tocart = data_tocart.replace("|", ",").replace("'", '"')
    #
    #     record = model_categories.search([("id", "=", categories_id)])
    #
    #     record.img_url_analysis_data_this_year = url
    #     record.img_data_analysis_data_this_year = {
    #         "hits_view": json.loads(data_hits),
    #         "hits_tocart": json.loads(data_tocart),
    #     }

    # def import_images_categorie_categorie_sale_this_year(self, content):
    #     model_categories = self.env["ozon.categories"]
    #
    #     model, categories_id, url, average_data = content.split(",")
    #     average_data = average_data.replace("|", ",")
    #
    #     record = model_categories.search([("id", "=", categories_id)])
    #
    #     record.img_url_sale_this_year = url
    #     record.img_data_sale_this_year = average_data

    # def import_images_categorie_categorie_sale_last_year(self, content):
    #     model_categories = self.env["ozon.categories"]
    #
    #     model, categories_id, url, average_data = content.split(",")
    #     average_data = average_data.replace("|", ",")
    #
    #     record = model_categories.search([("id", "=", categories_id)])
    #
    #     record.img_url_sale_last_year = url
    #     record.img_data_sale_last_year = average_data

    def import_actions(self, content) -> dict:
        qty_actions = 0
        qty_created_actions = 0
        qty_created_action_candidate = 0
        qty_action_participants = 0
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            for i, row in enumerate(reader):
                qty_actions += 1
                a_id = row["action_id"]
                is_participating = ast.literal_eval(row["is_participating"])
                potential_prod_count = row["potential_products_count"]
                participating_products_count = row["participating_products_count"]
                if action := self.env["ozon.action"].search([("a_id", "=", a_id)]):
                    action.write(
                        {
                            "is_participating": is_participating,
                            "participating_products_count": participating_products_count,
                        }
                    )
                else:
                    datetime_start = convert_ozon_datetime_str_to_odoo_datetime_str(
                        row["date_start"]
                    )
                    datetime_end = convert_ozon_datetime_str_to_odoo_datetime_str(
                        row["date_end"]
                    )
                    action = self.env["ozon.action"].create(
                        {
                            "a_id": a_id,
                            "name": row["name"],
                            "with_targeting": row["with_targeting"],
                            "datetime_start": datetime_start,
                            "datetime_end": datetime_end,
                            "description": row["description"],
                            "action_type": row["action_type"],
                            "discount_type": row["discount_type"],
                            "discount_value": row["discount_value"],
                            "potential_products_count": potential_prod_count,
                            "is_participating": is_participating,
                            "participating_products_count": participating_products_count,
                        }
                    )
                    qty_created_actions += 1

                    candidates = json.loads(row["action_candidates"])
                    candidates_data = []
                    len_candidates = len(candidates)
                    ids_on_platform = []
                    for idx, can in enumerate(candidates):
                        id_on_platform = can["id"]
                        if ozon_product := self.is_ozon_product_exists(id_on_platform):
                            candidates_data.append(
                                {
                                    "action_id": action.id,
                                    "product_id": ozon_product.id,
                                    "id_on_platform": id_on_platform,
                                    "max_action_price": can["max_action_price"],
                                }
                            )
                            ids_on_platform.append(id_on_platform)
                            print(
                                f"{idx}/{len_candidates} - Product {id_on_platform} was added as action {a_id} candidate"
                            )
                            qty_created_action_candidate += 1

                    action_candidate_ids = (
                        self.env["ozon.action_candidate"].create(candidates_data).ids
                    )
                    action.write(
                        {
                            "action_candidate_ids": action_candidate_ids,
                            "ids_on_platform": ids_on_platform,
                        }
                    )

                if is_participating:
                    participants = json.loads(row["action_participants"])
                    for par in participants:
                        id_on_platform = str(par["id_on_platform"])
                        action_candidate = action.action_candidate_ids.filtered(
                            lambda r: r.product_id_on_platform == id_on_platform
                        )
                        if action_candidate:
                            action_candidate.is_participating = True
                            qty_action_participants += 1

                # print(f"{i} - Action {a_id} was imported")

        log_data = {
            "Всего получено данных об акциях": qty_actions,
            "Создано новых акций": qty_created_actions,
            "Создано новых кандидатов к акциям": qty_created_action_candidate,
            "Добавлено в участники акций": qty_action_participants,
        }

        return log_data

    def import_successful_products_competitors(self, content):
        lines = content.split("\n")

        for line in lines[1:]:
            if not line:
                continue

            sku, name = line.split(",")

            model_successful_products_competitors = self.env[
                "ozon.successful_product_competitors"
            ]
            model_successful_products_competitors.create(
                {
                    "sku": sku,
                    "name": name,
                }
            )

    def _parse_ozon_competitors_goods_xlsx_file(self, wb):
        for sheet in wb.worksheets:
            period_from = None
            period_to = None
            category = None

            for row in sheet.iter_rows(min_row=1, max_row=4, values_only=True):
                if row[0]:
                    if "Период:" in row[0]:
                        splited_row = row[0].split(" ")
                        for str_ in splited_row:
                            if str_[0].isnumeric() and not period_from:
                                period_from = date(*[int(x) for x in str_.split("-")])
                                continue
                            if str_[0].isnumeric() and period_from and not period_to:
                                period_to = date(*[int(x) for x in str_.split("-")])
                                break
                    elif "Категория:" in row[0]:
                        splited_row = row[0].split(": ")
                        category = splited_row[1]

            existed_category = self.env["ozon.categories"].search(
                [("name_categories", "=", category)]
            )
            report = self.env["ozon.report_category_market_share"].create(
                {
                    "ozon_categories_id": existed_category.id if existed_category else False,
                    "period_from": period_from,
                    "period_to": period_to,
                }
            )

            # accumulate data from file
            sellers_cache = defaultdict(lambda: None)
            competitors_sales_ids = []
            sales_and_share_per_seller = defaultdict(int)
            for row in sheet.iter_rows(
                min_row=4, max_col=8, max_row=sheet.max_row, values_only=True
            ):
                if row[0] and row[0] != "№":
                    competitor_name = row[1]
                    product_name = row[2]
                    article = row[3]
                    category_lvl2 = row[4]
                    category_lvl3 = row[5]
                    ordered_quantity = row[6]
                    ordered_amount = row[7]

                    # get seller
                    seller = sellers_cache.get(competitor_name)
                    if not seller:
                        seller = self._get_seller(competitor_name)
                        sellers_cache[competitor_name] = seller

                    product_competitor_or_product = (
                        self._get_product_competitor_or_product(article, seller)
                    )
                    sales_and_share_per_seller[seller] += ordered_amount

                    competitor_sale = self._get_competitor_sale(
                        product_competitor_or_product,
                        period_from,
                        period_to,
                        ordered_quantity,
                        ordered_amount,
                        category_lvl3,
                        seller,
                        product_name,
                        article,
                    )
                    competitors_sales_ids.append(competitor_sale.id)

            # create ozon.report.competitor_category_share
            sellers = [
                (seller, shares)
                for seller, shares in sales_and_share_per_seller.items()
            ]
            sellers.sort(key=lambda x: x[1], reverse=True)
            place = 1
            competitors_category_share_ids = []
            for seller_tuple in sellers:
                competitor_category_share = self.env[
                    "ozon.report.competitor_category_share"
                ].create(
                    {
                        "place": place,
                        "seller_name": seller_tuple[0].trade_name,
                        "turnover": seller_tuple[1],
                    }
                )
                competitors_category_share_ids.append(competitor_category_share.id)
                place += 1

            # add data to report
            report.ozon_products_competitors_sale_ids = competitors_sales_ids
            report.ozon_report_competitor_category_share_ids = (
                competitors_category_share_ids
            )

        wb.close()

    def _get_competitor_sale(
        self,
        product_competitor_or_product,
        period_from,
        period_to,
        ordered_quantity,
        ordered_amount,
        category_lvl3,
        seller,
        product_name,
        article,
    ):
        competitor_sale = None
        if not seller.is_my_shop:
            if product_competitor_or_product:
                for (
                    sale
                ) in product_competitor_or_product.ozon_products_competitors_sale_ids:
                    if (
                        sale.period_from == period_from
                        and sale.period_to == period_to
                        and sale.seller_name == seller.trade_name
                    ):
                        sale.orders_qty = ordered_quantity
                        sale.orders_sum = ordered_amount
                        competitor_sale = sale
            if not competitor_sale:
                competitor_sale = self.env["ozon.products_competitors.sale"].create(
                    {
                        "period_from": period_from,
                        "period_to": period_to,
                        "ozon_products_competitors_id": (
                            product_competitor_or_product.id
                            if product_competitor_or_product
                            else False
                        ),
                        "orders_qty": ordered_quantity,
                        "orders_sum": ordered_amount,
                        "category_lvl3": category_lvl3,
                        "seller_name": seller.trade_name,
                        "name": f"{article}, {product_name}",
                    }
                )
        else:
            competitor_sale = self.env["ozon.products_competitors.sale"].create(
                {
                    "period_from": period_from,
                    "period_to": period_to,
                    "ozon_products_id": product_competitor_or_product.id,
                    "orders_qty": ordered_quantity,
                    "orders_sum": ordered_amount,
                    "category_lvl3": category_lvl3,
                    "seller_name": seller.trade_name,
                    "name": f"{article}, {product_name}",
                }
            )

        return competitor_sale

    def _get_product_competitor_or_product(self, article, seller) -> Any:
        if seller.is_my_shop:
            product = self.env["ozon.products"].search(
                [
                    ("article", "=", article),
                    # ('seller', '=', seller.id),
                ],
                limit=1,
            )
            if not product:
                pass
            return product
        else:
            product_competitor = self.env["ozon.products_competitors"]\
                .search(
                [
                    ("article", "=", article),
                    ("competitor_seller_id", "=", seller.id),
                ],
                limit=1,
            )

            return product_competitor

    def _get_seller(self, competitor_name: str) -> Any:
        seller = self.env["ozon.competitor_seller"].search(
            [("trade_name", "=", competitor_name)], limit=1
        )
        if not seller:
            seller = self.env["retail.seller"].search(
                [("trade_name", "=", competitor_name)], limit=1
            )
        if not seller:
            seller = self.env["ozon.competitor_seller"].create(
                {"trade_name": competitor_name}
            )

        return seller


class ProcessProductFile(models.Model):
    _inherit = "ozon.import_file"

    def _get_all_categories(self) -> dict:
        cats = self.env['ozon.categories'].search([])
        res = {}
        for cat in cats:
            res[cat.name_categories] = cat.id
        return res

    def _get_or_create_category(self, name_categories, c_id):
        model = self.env["ozon.categories"]
        record = model.search([("c_id", "=", c_id)])
        if not record:
            record = model.create({
                "name_categories": name_categories,
                "c_id": c_id,
            })
        return record

    def get_or_create_seller(self, seller_name: str):
        model_seller = self.env["retail.seller"]
        seller = model_seller.search([("name", "=", seller_name)], limit=1)
        if not seller:
            seller = model_seller.create({
                "name": "Продавец",
                "trade_name": "mobparts",
            })
        return seller

    def populate_supplementary_categories(
        self, full_categories_string: str, full_categories_id: int
    ):
        cats_list = split_keywords_on_slash(full_categories_string)
        cats_list = remove_duplicates_from_list(cats_list)

        sup_cat_vals = []
        for cat in cats_list:
            if not self.env["ozon.supplementary_categories"].search([
                    ("sc_id", "=", full_categories_id), ("name", "=", cat)
            ]):
                vals = {"sc_id": int(full_categories_id), "name": cat}
                sup_cat_vals.append(vals)

        return sup_cat_vals

    def _create_supplementary_categories(self, sup_categories: dict) -> list:
        vals = []
        model = self.env["ozon.supplementary_categories"]
        for full_categories_id, full_categories in sup_categories.items():
            supplementary_categories_vals = self.populate_supplementary_categories(
               full_categories, int(full_categories_id)
            )
            vals.extend(supplementary_categories_vals)
        model.create(vals)
        all_sup_cats = model.search([])
        res = [{'full_categories_id': str(cat.sc_id), 'id': cat.id, 'name': cat.name} for cat in all_sup_cats]
        return res

    def populate_search_queries(self, keywords_string: str, ozon_product_id) -> list:
        keywords = split_keywords(keywords_string)
        vals = []
        for word in keywords:
            record = self.env["ozon.search_queries"].search([
                ("words", "=", word),
                ("product_id", "=", ozon_product_id),
            ])
            if record:
                continue

            vals.append({"words": word, "product_id": ozon_product_id})
        return vals

    def process_products_imported_data(self, content) -> dict:
        with StringIO(content) as csvfile:
            reader = csv.DictReader(csvfile)

            ids_on_platform = []
            imported_products_vals = {}
            imported_categories = {}
            retail_products_data = {}
            retail_products_offer_ids = []
            sup_categories = {}
            data_qty = 0
            for old_row in reader:
                # products
                id_on_platform = old_row.get('id_on_platform')
                ids_on_platform.append(id_on_platform)
                imported_products_vals[id_on_platform] = old_row
                # categories
                c_id = int(old_row.get('description_category_id'))
                if c_id and not imported_categories.get(c_id):
                    imported_categories[c_id] = {
                        'name_categories': old_row.get('categories'),
                        'c_id': c_id,
                    }
                # retail products
                retail_products_offer_ids.append(old_row.get('offer_id'))
                if not retail_products_data.get(old_row.get('offer_id')):
                    retail_products_data[old_row.get('offer_id')] = {
                        'product_id': old_row.get('offer_id'),
                        'name': old_row.get('name'),
                        'description': old_row.get('description'),
                        'length': float(old_row.get('length')),
                        'width': float(old_row.get('width')),
                        'height': float(old_row.get('height')),
                        'weight': float(old_row.get('weight')),
                        'keywords': old_row.get('keywords') if old_row.get('keywords') else '',
                    }
                full_categories_id = old_row['full_categories_id']
                if not sup_categories.get(full_categories_id):
                    if full_categories_id:
                        sup_categories[full_categories_id] = old_row['full_categories']

                data_qty += 1

        # update cats data
        updated_cats_qty, created_cats_qty = self._create_update_categories(imported_categories)
        del imported_categories

        # update retail_products_data
        retail_products_dict = self._get_retail_products_dict(retail_products_offer_ids)
        curr_retail_products = self._create_update_retail_products_and_get_ids(
            retail_products_dict, retail_products_data, retail_products_offer_ids
        )
        del retail_products_offer_ids
        del retail_products_dict
        del retail_products_data

        # get data
        curr_categories = self._get_all_categories()

        # update ozon products
        products_dict = self._get_products_dict(ids_on_platform)
        updated_products_qty, created_products_qty = self._create_update_products(
            imported_products_vals, products_dict, curr_categories, curr_retail_products)
        del products_dict
        del curr_categories
        del curr_retail_products

        # other
        all_sup_categories = self._create_supplementary_categories(sup_categories)
        del sup_categories

        products_dict = self._get_products_dict(ids_on_platform)
        res = self._create_search_queries_and_get_sup_categories_ids(
            imported_products_vals, products_dict, all_sup_categories
        )
        products_ids, sup_categories_ids = res
        processed_products_qty, qty_new_price_history = self._write_sup_categories_ids_fees_fix_expenses_price_history(
            products_ids, sup_categories_ids, imported_products_vals
        )
        logger.warning(f"{processed_products_qty} products processed")

        log_data = {
            'Всего получено данных о продуктах': data_qty,
            'Обработано продуктов': processed_products_qty,
            'Обновлено продуктов': updated_products_qty,
            'Создано новых историй цен': qty_new_price_history,
            'Создано продуктов': created_products_qty,
            'Обновлено категорий': updated_cats_qty,
            'Создано категорий': updated_cats_qty,
        }
        return log_data

    def _create_search_queries_and_get_sup_categories_ids(
            self, imported_products_vals, products_dict, all_sup_categories
    ) -> tuple:
        search_queries_vals_to_create = []
        products_ids_with_sup_categories_ids = defaultdict(list)
        products_ids = []
        for id_on_platform, data in imported_products_vals.items():
            curr_product_data = products_dict.get(id_on_platform)
            # check and update if necessary
            if curr_product_data:
                product_id = curr_product_data['id']
                products_ids.append(product_id)
                # create search queries vals
                keywords = data['keywords']
                if keywords:
                    search_queries_vals = self.populate_search_queries(keywords, curr_product_data['id'])
                    search_queries_vals_to_create.extend(search_queries_vals)
                # sup categories get ids
                full_categories = data['full_categories']
                full_categories_id = data['full_categories_id']
                for cat in all_sup_categories:
                    if cat['full_categories_id'] == full_categories_id:
                        if cat['name'] in full_categories:
                            cat_id = cat['id']
                            products_ids_with_sup_categories_ids[product_id].append(cat_id)

        self.env["ozon.search_queries"].create(search_queries_vals_to_create)

        return products_ids, products_ids_with_sup_categories_ids

    def _write_sup_categories_ids_fees_fix_expenses_price_history(
            self, products_ids, products_ids_with_sup_categories_ids, imported_products_vals) -> tuple:
        qty_new_price_history = 0
        # sup categories check and write ids
        ozon_products = self.env['ozon.products'].browse(products_ids)
        price_histories_vals_to_write = []
        for ozon_product in ozon_products:
            sup_cats_ids = products_ids_with_sup_categories_ids.get(ozon_product.id)
            if sup_cats_ids:
                if ozon_product.supplementary_categories.ids != sup_cats_ids:
                    ozon_product.supplementary_categories = sup_cats_ids

            # fees
            id_on_platform = ozon_product.id_on_platform
            old_row = imported_products_vals[id_on_platform]
            all_fees = {k: old_row[k] for k in ALL_COMMISSIONS.keys()}

            if product_fee := ozon_product.product_fee:
                if product_fee.product_id_on_platform != id_on_platform:
                    product_fee.write({
                        "product_id_on_platform": id_on_platform
                    })
                are_fees_the_same = True
                for key, new_value in all_fees.items():
                    if product_fee[key] != float(new_value):
                        are_fees_the_same = False
                        product_fee.write({
                            "product_id_on_platform": id_on_platform,
                            **all_fees,
                        })
                        break
            else:
                are_fees_the_same = False
                product_fee = self.env["ozon.product_fee"].create({
                    "product": ozon_product.id,
                    "product_id_on_platform": id_on_platform,
                    **all_fees,
                })
                ozon_product.write({"product_fee": product_fee.id})

            # fix expenses
            if are_fees_the_same:
                fix_expenses_ids = ozon_product.fix_expenses.ids
                percent_expenses_ids = ozon_product.percent_expenses.ids

            else:
                fix_expenses = self.env["ozon.fix_expenses"].create_from_ozon_product_fee(product_fee)
                fix_expenses_ids = fix_expenses.ids

                percent_expenses = self.env["ozon.cost"].create_from_ozon_product_fee(
                    product_fee=product_fee,
                    price=ozon_product.price,
                )
                percent_expenses_ids = percent_expenses.ids

                ozon_product.write(
                    {
                        "fix_expenses": fix_expenses_ids,
                        "percent_expenses": percent_expenses_ids,
                    },
                    percent_expenses=percent_expenses,
                )

            # price history
            previous_price_history = self.env["ozon.price_history"].search([
                ('product', '=', ozon_product.id)
            ], order="create_date desc", limit=1)
            previous_price = previous_price_history.price if previous_price_history else 0
            previous_marketing_price = previous_price_history.marketing_price if previous_price_history else 0

            if (
                    previous_price != ozon_product.price or
                    previous_marketing_price != ozon_product.marketing_price
            ):
                price_history_data = {
                    "product": ozon_product.id,
                    "id_on_platform": id_on_platform,
                    "provider": ozon_product.seller.id,
                    "price": ozon_product.price,
                    "marketing_price": ozon_product.marketing_price,
                    "previous_price": previous_price,
                    "previous_marketing_price": previous_marketing_price,
                    "fix_expenses": fix_expenses_ids,
                    "costs": percent_expenses_ids,
                }
                price_histories_vals_to_write.append(price_history_data)
                qty_new_price_history += 1
        self.env["ozon.price_history"].create(price_histories_vals_to_write)

        return len(ozon_products), qty_new_price_history

    def _create_update_products(
            self, imported_products_vals, products_dict, curr_categories, curr_retail_products) -> tuple:
        vals_to_create_products = []
        updated_products_qty = 0
        created_products_qty = 0
        for id_on_platform, data in imported_products_vals.items():
            curr_product_data = products_dict.get(id_on_platform)
            # check and update if necessary
            if curr_product_data:
                vals = {}
                for key, curr_val in curr_product_data.items():
                    if key == 'imgs_urls':
                        key = 'img_urls'
                    elif key == 'article':
                        key = 'offer_id'
                    elif key == 'id':
                        continue

                    new_val = data[key] if key != 'products' and key != 'seller' else None
                    if key == 'marketing_price':            
                        new_val = float(new_val) if new_val else 0
                    elif key == 'price':
                        new_val = float(new_val) if new_val else 0
                    elif key == 'old_price':
                        new_val = float(new_val) if new_val else 0
                    elif key == 'ext_comp_min_price':
                        new_val = float(new_val) if new_val else 0
                    elif key == 'ozon_comp_min_price':
                        new_val = float(new_val) if new_val else 0
                    elif key == 'self_marketplaces_min_price':
                        new_val = float(new_val) if new_val else 0
                    elif key == 'categories':
                        if new_val:
                            category_id = curr_categories.get(new_val)
                            new_val = category_id
                    elif key == 'products':
                        offer_id = data.get('offer_id')
                        retail_prod_id = curr_retail_products.get(offer_id)
                        new_val = retail_prod_id
                    elif key == 'seller':
                        if not curr_val:
                            seller_id = self.get_or_create_seller(data['seller_name']).id
                            vals[key] = seller_id
                            continue
                        continue

                    if curr_val != new_val:
                        # logger.warning(f"{key}{type(curr_val)}")
                        # logger.warning(f"{key}{type(new_val)}")
                        # logger.warning('-----------------')
                        if key == 'img_urls':
                            key = 'imgs_urls'
                        elif key == 'offer_id':
                            key = 'article'
                        vals[key] = new_val
                if vals:
                    product = self.env['ozon.products'].search([("id_on_platform", '=', id_on_platform)])
                    product.sudo().write(vals)
                    updated_products_qty += 1
            # create
            else:
                # get category
                name_categories = data.get('categories')
                category_id = curr_categories.get(name_categories)
                # get retail product
                offer_id = data.get('offer_id')
                retail_prod_id = curr_retail_products.get(offer_id)
                seller = self.get_or_create_seller(data['seller_name'])
                vals = {
                    "id_on_platform": data['id_on_platform'],
                    "sku": data["sku"],
                    "fbo_sku": data["fbo_sku"],
                    "fbs_sku": data["fbs_sku"],
                    "categories": category_id,
                    "article": data['offer_id'],
                    "description": data['description'],
                    "products": retail_prod_id,
                    "marketing_price": data['marketing_price'],
                    "price": data['price'],
                    "old_price": data['old_price'],
                    "ext_comp_min_price": data["ext_comp_min_price"],
                    "ozon_comp_min_price": data["ozon_comp_min_price"],
                    "self_marketplaces_min_price": data[
                        "self_marketplaces_min_price"
                    ],
                    "price_index": data["price_index"],
                    "imgs_urls": data["img_urls"],
                    "seller": seller.id,
                    "trading_scheme": data["trading_scheme"],
                }
                vals_to_create_products.append(vals)
                created_products_qty += 1

        self.env['ozon.products'].create(vals_to_create_products)

        return updated_products_qty, created_products_qty

    def _create_update_categories(self, imported_cats_data: dict):
        updated_cats_qty = 0
        created_cats_qty = 0
        model = self.env["ozon.categories"]
        curr_categories = model.search([])
        curr_cats_data = {
            cat.c_id: {'name_categories': cat.name_categories, 'c_id': cat.c_id} for cat in curr_categories
        }
        for c_id, data in imported_cats_data.items():
            curr_cat_vals = curr_cats_data.get(c_id)
            if curr_cat_vals:
                if data != curr_cat_vals:
                    category = model.search([('c_id', '=', c_id)], limit=1)
                    category.write(data)
                    updated_cats_qty += 1
            else:
                model.create(data)
                created_cats_qty += 1

        return updated_cats_qty, created_cats_qty

    def _get_products_dict(self, ids_on_platform) -> dict:
        query = """
                SELECT
                id_on_platform,
                sku,
                fbo_sku,
                fbs_sku,
                categories,
                article,
                description,
                products,
                marketing_price,
                price,
                old_price,
                ext_comp_min_price,
                ozon_comp_min_price,
                self_marketplaces_min_price,
                price_index,
                imgs_urls,
                seller,
                trading_scheme,
                id
                FROM ozon_products
                WHERE id_on_platform IN %s
                """
        self.env.cr.execute(query, (tuple(ids_on_platform),))
        products_raw_vals = self.env.cr.fetchall()

        products_dict = {}
        for product_vals in products_raw_vals:
            id_on_platform = product_vals[0]
            if id_on_platform:
                product_data = {
                    "sku": product_vals[1],
                    "fbo_sku": product_vals[2],
                    "fbs_sku": product_vals[3],
                    "categories": product_vals[4],
                    "article": product_vals[5],
                    "description": product_vals[6],
                    "products": product_vals[7],
                    "marketing_price": product_vals[8],
                    "price": product_vals[9],
                    "old_price": product_vals[10],
                    "ext_comp_min_price": product_vals[11],
                    "ozon_comp_min_price": product_vals[12],
                    "self_marketplaces_min_price": product_vals[13],
                    "price_index": product_vals[14],
                    "imgs_urls": product_vals[15],
                    "seller": product_vals[16],
                    "trading_scheme": product_vals[17],
                    "id": product_vals[18],
                }
                products_dict[id_on_platform] = product_data

        return products_dict

    def _create_update_retail_products_and_get_ids(
            self, curr_retail_products, new_retail_products, offer_ids
    ) -> dict:
        create_data = []
        model = self.env['retail.products']
        for product_id, new_data in new_retail_products.items():
            curr_data = curr_retail_products.get(product_id)
            if curr_data:
                if curr_data != new_data:
                    vals = {}
                    for key, value in new_data.items():
                        if new_data[key] != curr_data[key]:
                            vals[key] = new_data[key]
                    ret_product = model.search([('product_id', '=', product_id)])
                    if ret_product:
                        ret_product.write(vals)
                    else:
                        create_data.append(new_data)
            else:
                create_data.append(new_data)
        # create
        model.create(create_data)

        # update
        query = """
                SELECT
                    product_id,
                    id
                FROM retail_products
                WHERE product_id IN %s
                """
        self.env.cr.execute(query, (tuple(offer_ids),))
        products_raw_vals = self.env.cr.fetchall()
        retail_products_dict = {product[0]: product[1] for product in products_raw_vals}

        return retail_products_dict

    def _get_retail_products_dict(self, offer_ids: list) -> dict:
        query = """
                SELECT
                    product_id,
                    name,
                    description,
                    length,
                    width,
                    height,
                    weight,
                    keywords
                FROM retail_products
                WHERE product_id IN %s
                """
        self.env.cr.execute(query, (tuple(offer_ids),))
        products_raw_vals = self.env.cr.fetchall()
        retail_products_dict = {}
        for product_vals in products_raw_vals:
            product_id_offer_id_article = product_vals[0]
            vals = {
                'product_id': product_id_offer_id_article,
                'name': product_vals[1],
                'description': product_vals[2],
                'length': product_vals[3],
                'width': product_vals[4],
                'height': product_vals[5],
                'weight': product_vals[6],
                'keywords': product_vals[7],
            }
            retail_products_dict[product_id_offer_id_article] = vals

        return retail_products_dict


old_row = {'id_on_platform': '85658502',
           'offer_id': '020339',
           'sku': '279467165',
           'fbo_sku': '0',
           'fbs_sku': '0',
           'categories': 'Дисплеи для телефонов',
           'description_category_id': '53567477',
           'full_categories': 'Miscellaneous accessories/Запчасти Access/Запчасти для телефонов Access/Дисплеи для телефонов',
           'full_categories_id': '971463128',
           'name': 'Модуль (матрица + тачскрин) для Huawei Honor, Y5 II 3G (CUN-U29) золотой',
           'description': 'Модуль  для Huawei Honor 6C / Enjoy 6S белый',
           'keywords': '',
           'length': '0.7',
           'width': '0.7',
           'height': '0.5',
           'weight': '0.06',
           'seller_name': 'Продавец',
           'trading_scheme': 'undefined',
           'price': '1367.0000',
           'old_price': '2367.0000',
           'ext_comp_min_price': '',
           'ozon_comp_min_price': '',
           'self_marketplaces_min_price': '',
           'price_index': 'WITHOUT_INDEX',
           'acquiring': '18',
           'fbo_fulfillment_amount': '0',
           'fbo_direct_flow_trans_min_amount': '0',
           'fbo_direct_flow_trans_max_amount': '0',
           'fbo_deliv_to_customer_amount': '75.19',
           'fbo_return_flow_amount': '0',
           'fbo_return_flow_trans_min_amount': '63',
           'fbo_return_flow_trans_max_amount': '63',
           'fbs_first_mile_min_amount': '0',
           'fbs_first_mile_max_amount': '25',
           'fbs_direct_flow_trans_min_amount': '76',
           'fbs_direct_flow_trans_max_amount': '76',
           'fbs_deliv_to_customer_amount': '75.19',
           'fbs_return_flow_amount': '0',
           'fbs_return_flow_trans_min_amount': '76',
           'fbs_return_flow_trans_max_amount': '76',
           'sales_percent_fbo': '21.5', 'sales_percent_fbs': '22.5',
           'sales_percent': '22.5',
           'img_urls': "['https://cdn1.ozone.ru/s3/multimedia-w/6070605032.jpg']"}
